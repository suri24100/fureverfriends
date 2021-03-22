from openpyxl import Workbook, load_workbook

file_name1 = 'petfinder_description.xlsx'
file_name2 = 'petfinder-profiledata-allanimals.xlsx'

workbook1 = load_workbook(file_name1)
workbook2 = load_workbook(file_name2)

sheet1 = workbook1['petfinder_description']
sheet2 = workbook2['Sheet 1 - petfinder-profiledata']

for row in sheet1.iter_rows(min_row=2,max_row=sheet1.max_row):
    for row2 in sheet2.iter_rows(min_row=3,max_row=sheet2.max_row, min_col=1, max_col=4):
        if row[0].value == row2[0].value:
            sheet2.cell(row2[0].row, 4, row[1].value)

workbook2.save(file_name2)