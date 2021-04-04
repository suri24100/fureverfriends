from openpyxl import Workbook, load_workbook
import operator
import datetime

file_name1 = 'petfinder-profiledata-allanimals_CLEAN.xlsx'

workbook1 = load_workbook(file_name1)

sheet1 = workbook1['Sheet 1 - petfinder-profiledata']
keywords = ["playful", "playing", "play", "loves playing", "loves to play", "talkative", "adventurous",
            "affectionate", "independent","energetic","calm","protective", "quiet", "social"]

count = 3
for row in sheet1.iter_rows(min_row=3,max_row=5500, min_col=4, max_col=47):
    for cell in row:
        if isinstance(cell.value, int) or cell.value == None or isinstance(cell.value, datetime.datetime):
            continue
        elif any(word in cell.value.lower().replace(',', ' ').strip().split() for word in keywords):
            print(count)
            sheet1.delete_rows(cell.row,1)
            break
    count = count + 1

sheet1.delete_rows(5501,sheet1.max_row)


workbook1.save(file_name1)
