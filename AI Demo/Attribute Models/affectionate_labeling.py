from openpyxl import Workbook, load_workbook
import datetime
import operator

file_name1 = 'petfinder-profiledata-allanimals_AFFECTIONATE.xlsx'
#file_name1 = 'label_test.xlsx'
workbook1 = load_workbook(file_name1)

sheet1 = workbook1['Sheet 1 - petfinder-profiledata']
#sheet1 = workbook1['Sheet1']
affect = ["affectionate", "loving", "loving affectionate", "cuddle", "loves to cuddle",
          "sweet","gentle","sweet natured", "friendly", "super affectionate", "caring", "tender", "big heart",
          "big-hearted", "warmhearted", "cuddle", "cuddle-able"]
not_affect = ["not affectionate", "not sweet", "aggressive", "cold", "not cuddly", "unfeeling", "unloving",
              "aloof", "antisocial", "cold-eyed", "detached", "distant", "hard-hearted", "indifferent", "reserved",
              "unfeeling", "disaffected", "unfriendly", "not cuddling", "not like to cuddle", "hates cuddling",
              "hates affection", "not like to cuddle", "not like to be touched", "not like to be pet", "no petting"]

count = 0
for row in sheet1.iter_rows(min_row=3,max_row=sheet1.max_row + 1,min_col=4, max_col=47):
    for cell in row:
        if isinstance(cell.value, int) or cell.value == None or isinstance(cell.value, datetime.datetime):
            continue
        elif any(word in cell.value.lower().replace(',', ' ').strip().split() for word in not_affect) or any(word in cell.value.lower() for word in not_affect):
            sheet1.cell(cell.row,47, 0)
            break
        elif any(word in cell.value.lower().replace(',', ' ').strip().split() for word in affect) or any(word in cell.value.lower() for word in affect):
            if count < 20000:
                querywords = cell.value.split()
                resultwords = [word for word in querywords if word.lower().replace(',', ' ').strip() not in affect]
                result = ' '.join(resultwords)
                cell.value = result

            sheet1.cell(cell.row,47, 1)
            break

    count = count + 1
    print(count)

workbook1.save(file_name1)
