from openpyxl import Workbook, load_workbook
import datetime
import operator

file_name1 = 'petfinder-profiledata-allanimals_SOCIAL.xlsx'
#file_name1 = 'label_test.xlsx'
workbook1 = load_workbook(file_name1)

sheet1 = workbook1['Sheet 1 - petfinder-profiledata']
#sheet1 = workbook1['Sheet1']
social = ["social", "social butterfly", "talkative", "friendly", "social environments", "socialized" "outgoing",
          "social outgoing", "socializes", "extrovert", "affable", "communicative", "extraverted", "gregarious", "talks"]
not_social = ["undersocialized", "is aggressive", "antisocial", "cold-eyed", "detached", "distant", "unfriendly",
              "not social", "not friendly", "not talkative", "shy", "reserved", "mute", "silent" , "insociable",
              "introverted", "nongregarious", "reclusive", "unsociable", "unsocial", "quiet", "does not talk",
              "not like to talk", "unsociable", "not like other pets", "not like to be around other pets",
              "not like pets", "like to be alone"]

count = 0
for row in sheet1.iter_rows(min_row=3,max_row=sheet1.max_row + 1,min_col=4, max_col=47):
    for cell in row:
        if isinstance(cell.value, int) or cell.value == None or isinstance(cell.value, datetime.datetime):
            continue
        elif any(word in cell.value.lower().replace(',', ' ').strip().split() for word in not_social) or any(word in cell.value.lower() for word in not_social):
            sheet1.cell(cell.row,47, 0)
            break
        elif any(word in cell.value.lower().replace(',', ' ').strip().split() for word in social) or any(word in cell.value.lower() for word in social):
            if count < 20000:
                querywords = cell.value.split()
                resultwords = [word for word in querywords if word.lower().replace(',', ' ').strip() not in social]
                result = ' '.join(resultwords)
                cell.value = result

            sheet1.cell(cell.row,47, 1)
            break

    count = count + 1
    print(count)

workbook1.save(file_name1)