from openpyxl import Workbook, load_workbook
import operator

file_name1 = 'petfinder-profiledata-allanimals_FINAL.xlsx'


workbook1 = load_workbook(file_name1)

sheet1 = workbook1['Sheet 1 - petfinder-profiledata']

tag_count = {}

for row in sheet1.iter_rows(min_row=3,max_row=sheet1.max_row,min_col=5):
    for cell in row:
        if cell.value != "" and cell.value !=None:
            tag = cell.value.lower().strip()
            tag_count[tag] = tag_count.setdefault(tag, 0) + 1

sorted_tags = dict( sorted(tag_count.items(), key=operator.itemgetter(1),reverse=True))
for k , v in sorted_tags.items():
    print("{}: {}".format(k, v))
