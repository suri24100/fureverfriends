from openpyxl import Workbook, load_workbook
import datetime
import operator

file_name1 = 'petfinder-profiledata-allanimals_FINAL.xlsx'
#file_name1 = 'label_test.xlsx'
workbook1 = load_workbook(file_name1)

sheet1 = workbook1['Sheet 1 - petfinder-profiledata']
#sheet1 = workbook1['Sheet1']
playful = ["playful", "playing", "play", "loves playing", "loves to play", "likes to play", "energetic",
           "social", "active", "lively", "merry","spirited", "full of pep"]
not_playful = ["is serious", "calm", "lazy", "dispirited","depressed",
               "inactive", "lifeless", "unhappy", "shy", "couch potato"]

count = 0
for row in sheet1.iter_rows(min_row=3,max_row=sheet1.max_row + 1,min_col=4, max_col=47):
    for cell in row:
        if isinstance(cell.value, int) or cell.value == None or isinstance(cell.value, datetime.datetime):
            continue
        elif any(word in cell.value.lower().replace(',', ' ').strip().split() for word in playful):
            if count < 20000:
                querywords = cell.value.split()
                resultwords = [word for word in querywords if word.lower().replace(',', ' ').strip() not in playful]
                result = ' '.join(resultwords)
                cell.value = result

            sheet1.cell(cell.row,47, 1)
            break
        elif any(word in cell.value.lower().replace(',', ' ').strip().split() for word in not_playful):
            sheet1.cell(cell.row,47, 0)
            break

    count = count + 1
    print(count)

workbook1.save(file_name1)
